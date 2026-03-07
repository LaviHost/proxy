"""
Free Proxy Scraper — GitHub Actions Edition
============================================
Scrapes proxies from 6 public sources, checks every proxy,
and writes results to proxies.xlsx + summary.json.
GitHub Actions triggers this every hour via cron.
"""

import requests
import time
import re
import json
import logging
from datetime import datetime, timezone
from concurrent.futures import ThreadPoolExecutor, as_completed
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────
OUTPUT_XLSX   = "proxies.xlsx"
OUTPUT_JSON   = "summary.json"
CHECK_TIMEOUT = 8
MAX_WORKERS   = 60
TEST_URL      = "http://httpbin.org/ip"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0 Safari/537.36"
    )
}

# ── SCRAPERS ──────────────────────────────────────────────────────────────────

def scrape_free_proxy_list():
    proxies = []
    try:
        r = requests.get("https://free-proxy-list.net/", headers=HEADERS, timeout=15)
        soup = BeautifulSoup(r.text, "lxml")
        for row in soup.select("table tbody tr"):
            cols = [td.text.strip() for td in row.find_all("td")]
            if len(cols) >= 7:
                proxies.append({
                    "ip": cols[0], "port": cols[1],
                    "country_code": cols[2], "country": cols[3],
                    "anonymity": cols[4],
                    "https": "Yes" if cols[6] == "yes" else "No",
                    "protocol": "HTTPS" if cols[6] == "yes" else "HTTP",
                    "source": "free-proxy-list.net",
                })
        log.info(f"[free-proxy-list.net]   {len(proxies)}")
    except Exception as e:
        log.warning(f"[free-proxy-list.net] {e}")
    return proxies


def scrape_proxyscrape():
    proxies = []
    for proto in ["http", "socks4", "socks5"]:
        try:
            url = (
                f"https://api.proxyscrape.com/v2/?request=getproxies"
                f"&protocol={proto}&timeout=10000&country=all&ssl=all&anonymity=all"
            )
            r = requests.get(url, headers=HEADERS, timeout=15)
            for line in r.text.strip().splitlines():
                if ":" in line.strip():
                    ip, port = line.strip().split(":", 1)
                    proxies.append({
                        "ip": ip, "port": port,
                        "country_code": "", "country": "",
                        "anonymity": "unknown",
                        "https": "Yes" if proto == "http" else "No",
                        "protocol": proto.upper(),
                        "source": "proxyscrape.com",
                    })
        except Exception as e:
            log.warning(f"[proxyscrape/{proto}] {e}")
    log.info(f"[proxyscrape.com]       {len(proxies)}")
    return proxies


def scrape_proxynova():
    proxies = []
    try:
        r = requests.get(
            "https://www.proxynova.com/proxy-server-list/",
            headers=HEADERS, timeout=15
        )
        soup = BeautifulSoup(r.text, "lxml")
        for row in soup.select("table#tbl_proxy_list tbody tr"):
            cols = row.find_all("td")
            if len(cols) >= 6:
                abbr = cols[0].find("abbr")
                ip   = abbr["title"].strip() if abbr else ""
                port = cols[1].text.strip()
                if ip and port.isdigit():
                    proxies.append({
                        "ip": ip, "port": port,
                        "country_code": "", "country": cols[5].text.strip(),
                        "anonymity": cols[2].text.strip(),
                        "https": "", "protocol": "HTTP",
                        "source": "proxynova.com",
                    })
        log.info(f"[proxynova.com]         {len(proxies)}")
    except Exception as e:
        log.warning(f"[proxynova.com] {e}")
    return proxies


def scrape_geonode():
    proxies = []
    try:
        url = (
            "https://proxylist.geonode.com/api/proxy-list"
            "?limit=500&page=1&sort_by=lastChecked&sort_type=desc"
        )
        r = requests.get(url, headers=HEADERS, timeout=15)
        for item in r.json().get("data", []):
            proxies.append({
                "ip": item.get("ip", ""),
                "port": str(item.get("port", "")),
                "country_code": item.get("countryCode", ""),
                "country": item.get("country", ""),
                "anonymity": item.get("anonymityLevel", ""),
                "https": "Yes" if "https" in item.get("protocols", []) else "No",
                "protocol": "/".join(p.upper() for p in item.get("protocols", ["HTTP"])),
                "source": "geonode.com",
            })
        log.info(f"[geonode.com]           {len(proxies)}")
    except Exception as e:
        log.warning(f"[geonode.com] {e}")
    return proxies


def scrape_github_proxifly():
    proxies = []
    urls = {
        "HTTP":   "https://raw.githubusercontent.com/proxifly/free-proxy-list/main/proxies/protocols/http/data.txt",
        "SOCKS4": "https://raw.githubusercontent.com/proxifly/free-proxy-list/main/proxies/protocols/socks4/data.txt",
        "SOCKS5": "https://raw.githubusercontent.com/proxifly/free-proxy-list/main/proxies/protocols/socks5/data.txt",
    }
    for proto, url in urls.items():
        try:
            r = requests.get(url, headers=HEADERS, timeout=15)
            for line in r.text.strip().splitlines():
                line = re.sub(r"^https?://|^socks[45]://", "", line.strip())
                if ":" in line:
                    ip, port = line.rsplit(":", 1)
                    proxies.append({
                        "ip": ip, "port": port,
                        "country_code": "", "country": "",
                        "anonymity": "unknown",
                        "https": "Yes" if proto == "HTTP" else "No",
                        "protocol": proto, "source": "github/proxifly",
                    })
        except Exception as e:
            log.warning(f"[proxifly/{proto}] {e}")
    log.info(f"[github/proxifly]       {len(proxies)}")
    return proxies


def scrape_github_speedx():
    proxies = []
    urls = {
        "HTTP":   "https://raw.githubusercontent.com/TheSpeedX/PROXY-List/master/http.txt",
        "SOCKS4": "https://raw.githubusercontent.com/TheSpeedX/PROXY-List/master/socks4.txt",
        "SOCKS5": "https://raw.githubusercontent.com/TheSpeedX/PROXY-List/master/socks5.txt",
    }
    for proto, url in urls.items():
        try:
            r = requests.get(url, headers=HEADERS, timeout=15)
            for line in r.text.strip().splitlines():
                if ":" in line.strip():
                    ip, port = line.strip().split(":", 1)
                    proxies.append({
                        "ip": ip, "port": port,
                        "country_code": "", "country": "",
                        "anonymity": "unknown",
                        "https": "Yes" if proto == "HTTP" else "No",
                        "protocol": proto, "source": "github/TheSpeedX",
                    })
        except Exception as e:
            log.warning(f"[TheSpeedX/{proto}] {e}")
    log.info(f"[github/TheSpeedX]      {len(proxies)}")
    return proxies


# ── CHECKER ───────────────────────────────────────────────────────────────────

def check_proxy(proxy):
    proto     = proxy["protocol"].split("/")[0].lower()
    proto     = proto if proto in ("http", "https", "socks4", "socks5") else "http"
    proxy_url = f"{proto}://{proxy['ip']}:{proxy['port']}"
    start     = time.time()
    try:
        requests.get(
            TEST_URL,
            proxies={"http": proxy_url, "https": proxy_url},
            timeout=CHECK_TIMEOUT,
            headers=HEADERS,
        )
        proxy["status"]     = "✅ Live"
        proxy["latency_ms"] = round((time.time() - start) * 1000)
    except Exception:
        proxy["status"]     = "❌ Dead"
        proxy["latency_ms"] = None
    return proxy


def check_all(proxies):
    log.info(f"Checking {len(proxies)} proxies ({MAX_WORKERS} threads)...")
    results = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futures = {ex.submit(check_proxy, p): p for p in proxies}
        for i, fut in enumerate(as_completed(futures), 1):
            results.append(fut.result())
            if i % 500 == 0:
                log.info(f"  {i}/{len(proxies)} checked...")
    return results


# ── EXCEL WRITER ──────────────────────────────────────────────────────────────

C_NAVY    = "1F3864"
C_WHITE   = "FFFFFF"
C_GREEN   = "E2EFDA"
C_RED     = "FCE4D6"
C_GREY    = "F2F2F2"
C_BLUE    = "2E75B6"
C_YELLOW  = "FFF2CC"

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(ws, ref, val):
    c = ws[ref]
    c.value     = val
    c.font      = Font(bold=True, color=C_WHITE, name="Arial", size=10)
    c.fill      = PatternFill("solid", fgColor=C_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = _border()

COLS = [
    ("IP Address", 18), ("Port", 7), ("Protocol", 10), ("Country", 18),
    ("Code", 7), ("Anonymity", 15), ("HTTPS", 8),
    ("Status", 11), ("Latency (ms)", 13), ("Source", 22), ("Last Checked", 20),
]

def _apply_headers(ws):
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28
    for i, (name, width) in enumerate(COLS, 1):
        col = get_column_letter(i)
        _hdr(ws, f"{col}1", name)
        ws.column_dimensions[col].width = width
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

def _write_row(ws, row_idx, p, now_str, bg):
    fill = PatternFill("solid", fgColor=bg)
    vals = [
        p.get("ip"), p.get("port"), p.get("protocol"),
        p.get("country"), p.get("country_code"), p.get("anonymity"),
        p.get("https"), p.get("status"), p.get("latency_ms"),
        p.get("source"), now_str,
    ]
    for col_idx, val in enumerate(vals, 1):
        c = ws.cell(row=row_idx, column=col_idx, value=val)
        c.fill      = fill
        c.font      = Font(name="Arial", size=9)
        c.border    = _border()
        c.alignment = Alignment(horizontal="center", vertical="center")


def write_outputs(proxies):
    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    wb      = Workbook()

    # Sheet 1 — All
    ws_all = wb.active
    ws_all.title = "All Proxies"
    _apply_headers(ws_all)
    for i, p in enumerate(proxies, 2):
        live = p.get("status") == "✅ Live"
        bg   = C_GREEN if live else (C_RED if p.get("status") == "❌ Dead" else (C_GREY if i % 2 == 0 else C_WHITE))
        _write_row(ws_all, i, p, now_str, bg)

    # Sheet 2 — Live only, sorted by latency
    ws_live = wb.create_sheet("✅ Live Proxies")
    _apply_headers(ws_live)
    live_list = sorted(
        [p for p in proxies if p.get("status") == "✅ Live"],
        key=lambda x: x.get("latency_ms") or 9999
    )
    for i, p in enumerate(live_list, 2):
        bg = C_GREEN if i % 2 == 0 else "F0FFF0"
        _write_row(ws_live, i, p, now_str, bg)

    # Sheet 3 — Summary
    ws_sum = wb.create_sheet("📊 Summary")
    ws_sum.column_dimensions["A"].width = 26
    ws_sum.column_dimensions["B"].width = 20

    title = ws_sum["A1"]
    title.value = "🌐 Proxy Scraper — Summary"
    title.font  = Font(bold=True, size=13, color=C_BLUE, name="Arial")
    ws_sum.merge_cells("A1:B1")

    total      = len(proxies)
    live_count = len(live_list)
    sources    = {}
    protocols  = {}
    for p in proxies:
        sources[p.get("source", "?")] = sources.get(p.get("source", "?"), 0) + 1
        protocols[p.get("protocol", "?")] = protocols.get(p.get("protocol", "?"), 0) + 1

    avg_lat = ""
    lats = [p["latency_ms"] for p in live_list if p.get("latency_ms")]
    if lats:
        avg_lat = f"{round(sum(lats)/len(lats))} ms"

    rows = [
        ("Last Updated",    now_str,       C_YELLOW),
        ("Total Proxies",   total,         C_WHITE),
        ("✅ Live",          live_count,    C_GREEN),
        ("❌ Dead",          total-live_count, C_RED),
        ("Live Rate",       f"{round(live_count/total*100,1) if total else 0}%", C_WHITE),
        ("Avg Latency",     avg_lat,       C_WHITE),
        ("",                "",            C_WHITE),
        ("── Sources ──",   "Count",       C_GREY),
        *[(s, c, C_WHITE) for s, c in sorted(sources.items(), key=lambda x: -x[1])],
        ("",                "",            C_WHITE),
        ("── Protocols ──", "Count",       C_GREY),
        *[(p, c, C_WHITE) for p, c in sorted(protocols.items(), key=lambda x: -x[1])],
    ]
    for i, (label, value, bg) in enumerate(rows, 2):
        a = ws_sum.cell(row=i, column=1, value=label)
        b = ws_sum.cell(row=i, column=2, value=value)
        fill = PatternFill("solid", fgColor=bg)
        for c in (a, b):
            c.fill   = fill
            c.font   = Font(name="Arial", size=10, bold="──" in str(label))
            c.border = _border()
            c.alignment = Alignment(horizontal="left", vertical="center")

    wb.save(OUTPUT_XLSX)
    log.info(f"✅ Saved {OUTPUT_XLSX}")

    # JSON summary for badges / downstream use
    summary = {
        "last_updated": now_str,
        "total": total,
        "live": live_count,
        "dead": total - live_count,
        "live_rate_pct": round(live_count / total * 100, 1) if total else 0,
        "avg_latency_ms": round(sum(lats)/len(lats)) if lats else None,
        "by_source": sources,
        "by_protocol": protocols,
    }
    with open(OUTPUT_JSON, "w") as f:
        json.dump(summary, f, indent=2)
    log.info(f"✅ Saved {OUTPUT_JSON}")


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    log.info("=" * 55)
    log.info("  FREE PROXY SCRAPER  —  GitHub Actions run")
    log.info("=" * 55)

    all_proxies = []
    for scraper in [
        scrape_free_proxy_list, scrape_proxyscrape, scrape_proxynova,
        scrape_geonode, scrape_github_proxifly, scrape_github_speedx,
    ]:
        try:
            all_proxies.extend(scraper())
        except Exception as e:
            log.error(f"{scraper.__name__}: {e}")

    log.info(f"Raw total: {len(all_proxies)}")

    # Deduplicate
    seen, deduped = set(), []
    for p in all_proxies:
        key = f"{p['ip']}:{p['port']}"
        if key not in seen:
            seen.add(key)
            deduped.append(p)
    log.info(f"After dedup: {len(deduped)}")

    checked = check_all(deduped)
    live    = sum(1 for p in checked if p.get("status") == "✅ Live")
    log.info(f"Live: {live} / {len(checked)}")

    write_outputs(checked)


if __name__ == "__main__":
    main()
